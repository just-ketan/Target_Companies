import pandas as pd
import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill

# Input and output files
source_file = 'leetcode_analysis_categorized.xlsx'
output_file = 'leetcode_company_analysis_output.xlsx'

# Step 1: Read data
df = pd.read_excel(source_file)

# Step 2: Extract unique companies from all rows
companies = set()
for row in df['Companies'].dropna():
    for company in row.split(','):
        companies.add(company.strip())
companies = sorted(companies)

# Step 3: Function to scrape difficulty and topic from LeetCode (can be slow!)
def get_leetcode_info(link):
    try:
        page = requests.get(link)
        soup = BeautifulSoup(page.text, 'html.parser')
        # Scraping logic (update if LeetCode structure changes)
        difficulty = 'Unknown'
        topic = 'Unknown'
        for span in soup.find_all('span', class_='css-10d7fc9'):  # update selector as needed
            if span.text in ['Easy', 'Medium', 'Hard']:
                difficulty = span.text
        topic_tags = [a.text for a in soup.find_all('a', class_='topic-tag')]  # update selector as needed
        topic = topic_tags[0] if topic_tags else 'Unknown'
        return difficulty, topic
    except Exception:
        return 'Unknown', 'Unknown'

# Step 4: Build result rows with scraped info
result_rows = []
for idx, row in df.iterrows():
    for comp in str(row['Companies']).split(','):
        comp = comp.strip()
        if not comp: continue
        diff, topic = get_leetcode_info(row['Problem_Link'])
        result_rows.append({
            'Company': comp,
            'Problem Name': row['problem_name'],
            'LeetCode Link': row['Problem_Link'],
            'Difficulty': diff,
            'Topic': topic
        })

results_df = pd.DataFrame(result_rows)

# Step 5: Sort dataframe by difficulty mapping
diff_order = {'Easy': 0, 'Medium': 1, 'Hard': 2, 'Unknown': 3}
results_df['DiffOrder'] = results_df['Difficulty'].map(diff_order)
results_df = results_df.sort_values(by=['Company', 'DiffOrder'])

# Step 6: Write to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Company Filter'
ws.append(list(results_df.columns[:-1]))

# Step 7: Conditional formatting colors
diff_color = {'Easy': '00FF00', 'Medium': 'FFA500', 'Hard': 'FF0000', 'Unknown': 'CCCCCC'}

for r_idx, row in enumerate(results_df.values, 2):
    ws.append([row[0], row[1], row[2], row[3], row[4]])
    cell = ws.cell(row=r_idx, column=4)
    color = diff_color.get(row[3], 'CCCCCC')
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

# Step 8: Add dropdown for company selection
dv = DataValidation(type='list', formula1=f'"{",".join(companies)}"', allow_blank=True)
ws.add_data_validation(dv)
for row in range(2, ws.max_row + 1):
    dv.add(ws.cell(row=row, column=1))

# Step 9: Second sheet: Topic frequency for each company
topic_ws = wb.create_sheet('Topic Stats')
topic_ws.append(['Company', 'Topic', 'Frequency'])
for comp in companies:
    topics = results_df[results_df['Company'] == comp]['Topic'].value_counts()
    for topic, freq in topics.items():
        topic_ws.append([comp, topic, freq])
# Pie charts can be created using openpyxl or Excel itself later.

wb.save(output_file)
print('Excel file with dropdown, color coding & topic frequency exported!')
